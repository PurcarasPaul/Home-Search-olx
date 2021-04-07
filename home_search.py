import selenium.webdriver as webdriver
import xlsxwriter
import time
import openpyxl
import re
import ctypes
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

#function that gets older results if there are any
def get_old_results():
    try:
        temp_old_home_list_links = openpyxl.load_workbook('home_list.xlsx')
    except FileNotFoundError:
        return None
    else:
        sheet = temp_old_home_list_links.active
        temp_old_home_list_links = []
        for i in range(2,sheet.max_row+1):
             temp_old_home_list_links.append(sheet.cell(row=i,column=1).value)

        old_home_list_links = []
        for link in temp_old_home_list_links:
            old_home_list_links.append(link[0:link.index('html')+4])

        return old_home_list_links

#settings for chrome driver
def browser_options():
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/85.0.4183.83 Safari/537.36"

    options = webdriver.ChromeOptions()
    #setting that makes chrome invisible
    options.headless = True
    options.add_argument(f'user-agent={user_agent}')
    options.add_argument("--window-size=1920,1080")
    options.add_argument('--ignore-certificate-errors')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument("--disable-extensions")
    options.add_argument("--proxy-server='direct://'")
    options.add_argument("--proxy-bypass-list=*")
    options.add_argument("--start-maximized")
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--no-sandbox')

    return options

#eliminates duplicate links and links that contain "storia"(different site)
def isOk(result_link,results):
    if result_link.find('storia') != -1:
        return False
    if result_link in results:
        return False 
    return True
    
#checks if the links are good(no duplicate,no "storia") and returns a list with them
def get_links(links,results):
    for link in links:
        result_link=link.get_attribute("href")
        result_link = result_link[0:result_link.index('html')+4]
        if isOk(result_link,results):
            results.append(result_link)
    return results

#checks if there's a button for next page or not
def check_next_page(browser,xpath):
    try:
        next_page=browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return None
    else:
        return next_page.get_attribute("href")

#checks if the non-clickable button is an actual page or not
def isPage(browser,xpath):
    try:
        browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    else:
        return True

#returns the link to the next page if there is one
def next_page(browser):
    next_page_link = None
    element = 4
    converted_element = str(element)
    xpath = "//*[@id='body-container']/div[3]/div/div[4]/span["+converted_element+"]/a"
    while True:
        temp=check_next_page(browser,xpath)

        span_xpath=xpath[:-1]+"span"
        if temp != None or isPage(browser,span_xpath):
            next_page_link=temp
        else:
            break
            
        element+=1
        converted_element = str(element)
        xpath = "//*[@id='body-container']/div[3]/div/div[4]/span["+converted_element+"]/a"

    return next_page_link

#returns a list of final links
def get_results():
    url = input("Please enter a valid URL: ")
    browser = webdriver.Chrome('chromedriver',options=browser_options())
    browser.get(url)

    results = []
    while True:
        links=browser.find_elements_by_xpath("//h3//a")
        results=get_links(links,results)
        
        next_page_link = next_page(browser)
        if next_page_link != None:
            browser.get(next_page_link)
        else:
            break
        
    browser.close()
    return results

#checks if there is any forbidden word in text,if there is it will return false and the link will not be in excel
def isTextOk(text,forbidden_words):
    for word in forbidden_words:
        if re.search(r'\b' + word + r'\b', text) != None:
            return False
    return True

#returns text from the provided xpath,if there is any,otherwise it will return none
def get_data_xpath(temp_browser,xpath):
    try:
        temp_browser.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return None
    else:
        return temp_browser.find_element_by_xpath(xpath).text

#checks if the title and description contain any words that the user doesn't want there
def check_title_and_description(temp_browser,result_link,forbidden_words):
    #title
    text=get_data_xpath(temp_browser,"//div[1]/h1")
    if not text:
        text=get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/div[2]/h1")
    if not text:
        text = get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[1]/div[2]/h1")
    if isTextOk(text,forbidden_words)==False:
        return False

    #description
    text=get_data_xpath(temp_browser,"//*[@id='textContent']")
    if not text:
        text=get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/div[8]/div")
    if not text:
        text = get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[1]/div[8]/div")
    if isTextOk(text,forbidden_words)==False:
        return False

    return True

#checks 2 different xpaths to find the floor area and then returns it
def get_floor_area(temp_browser):
    floor_area_xpaths = ["//div[2]//ul//li[2]//span//strong","//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/ul/li[3]/p","//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/ul/li[2]/p"]

    floor_area=get_data_xpath(temp_browser,"//div[2]//ul//li[3]//span//strong")
    for xpath in floor_area_xpaths:
        if not floor_area:
            floor_area=get_data_xpath(temp_browser,xpath)

    if not floor_area:
        return None
    if "m²" in floor_area:
        return floor_area
    
#checks 2 different xpaths to find the floor and then returns it
def get_floor(temp_browser):
    floor_xpaths = ["//div[2]//ul//li[5]//a//strong","//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/ul/li[4]/p","//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/ul/li[5]/p"]

    floor = get_data_xpath(temp_browser,"//div[2]//ul//li[4]//a//strong")
    for xpath in floor_xpaths:
        if not floor or len(floor) > 6:
            floor = get_data_xpath(temp_browser,xpath)
    
    if not floor:
        return None
    return floor

#returns the date of when the offer was posted
def get_posted_date(temp_browser):
    posted_date = get_data_xpath(temp_browser,"//ul//li[1]//em//strong")
    
    if not posted_date:
        posted_date = get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/div[1]/span/span")
        return posted_date
    elif not posted_date:
        return None
    return posted_date[10:]

#clicks on a button to make the phone number visible then returns the phone number of the user that posted the offer,if there is any,otherwise it will return none
def get_phone_number(temp_browser):
    try:
        cookies = temp_browser.find_element_by_id("onetrust-accept-btn-handler")
    except NoSuchElementException:
        pass
    else:
        cookies.click()

    try:
        elem = temp_browser.find_element_by_xpath("//*[@id='contact_methods']/li[2]/div")
    except NoSuchElementException:
        try:
            elem = temp_browser.find_element_by_xpath("//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[4]/div/div/button")
        except NoSuchElementException:
            return None
        else:
            waitElement = WebDriverWait(temp_browser,10).until(EC.element_to_be_clickable((By.XPATH,("//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[4]/div/div/button"))))
            elem.click()
            waitPhone = WebDriverWait(temp_browser,10).until(EC.presence_of_element_located((By.XPATH,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[4]/div/div/ul/li")))
            phone_number = get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[4]/div/div/ul/li")
            return phone_number
    else:
        waitElement = WebDriverWait(temp_browser,10).until(EC.element_to_be_clickable((By.XPATH,("//*[@id='contact_methods']/li[2]/div"))))
        elem.click()
        waitPhone = WebDriverWait(temp_browser,10).until(EC.presence_of_element_located((By.XPATH,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[4]/div/div/ul/li")))
        phone_number = get_data_xpath(temp_browser,"//*[@id='contact_methods_below']/li/div/strong")
        return phone_number

#returns the price of the offer
def get_price(temp_browser):
    price = get_data_xpath(temp_browser,"//*[@id='offerdescription']/div[1]/div[2]/div/strong")

    if not price:
        price = get_data_xpath(temp_browser,"//*[@id='root']/div[1]/div[3]/div[2]/div[1]/div[2]/div[3]/h3")
    return price

#checks if the link is new comparing it to the older results of the same program
def new_link(result_link,old_results):
    if old_results != None and result_link in old_results:
        old_results.remove(result_link)
        return None
    return 'new'

#checks if the link is fine then gets multiple data from an offer,like price,phone number,etc.
def get_data(temp_browser,result_link,old_results,forbidden_words):
    temp_data = []
    temp_browser.get(result_link)

    if check_title_and_description(temp_browser,result_link,forbidden_words) != False:
        temp_data.append(result_link)
        temp_data.append(get_floor_area(temp_browser))
        temp_data.append(get_floor(temp_browser))
        temp_data.append(get_posted_date(temp_browser))
        temp_data.append(get_phone_number(temp_browser))
        temp_data.append(get_price(temp_browser))
        temp_data.append(new_link(result_link,old_results))

    return temp_data

#clicks a pop-up so that the other links won't have an issue while trying to click elements in them that would be covered by this pop-up
def accept_terms(temp_browser):
    temp_browser.get('https://www.olx.ro')
    waitCookies = WebDriverWait(temp_browser,10).until(EC.presence_of_element_located((By.ID,"onetrust-accept-btn-handler")))
    cookies = temp_browser.find_element_by_id("onetrust-accept-btn-handler").click()

#creates an excel file and fills it up with the data stored from each link
def results_to_excel(results,old_results,forbidden_words):
    temp_browser = webdriver.Chrome('chromedriver',options=browser_options())
    accept_terms(temp_browser)

    excel_file = xlsxwriter.Workbook("home_list.xlsx")
    sheet = excel_file.add_worksheet()
        
    sheet.write("A1","Link")
    sheet.write("B1","Floor Area")
    sheet.write("C1","Floor")
    sheet.write("D1","Posted")
    sheet.write("E1","Phone number")
    sheet.write("F1","Price")

    curent_link = 0
    for link in results:
        data=[]
        data.extend(get_data(temp_browser,link,old_results,forbidden_words))
            
        if data:
            curent_tab=0
            for single_data in data:
                sheet.write(curent_link+1,curent_tab,data[curent_tab])
                curent_tab+=1
            curent_link+=1

    temp_browser.close()
    excel_file.close()

old_results = get_old_results()
forbidden_words = input("Enter words with space between them that you don't want to be in the title or the description of the home:").split()
results = get_results()
results_to_excel(results,old_results,forbidden_words)
ctypes.windll.user32.MessageBoxW(0, "The program finished the task corectly!", "Success!")